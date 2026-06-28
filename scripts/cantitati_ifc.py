"""
Centralizator cantitati beton + cofraj din IFC
Workbook cu mai multe sheet-uri, format unitar printabil
Dependinte: pip install ifcopenshell openpyxl
"""

import sys, os, re, math
import ifcopenshell
import ifcopenshell.geom
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter

# ===========================================================================
# CONFIGURARE PROIECT — editeaza aici
# ===========================================================================
PROIECT = "DENUMIRE PROIECT"              # completeaza cu denumirea proiectului
ADRESA  = "Adresa proiectului"            # completeaza cu adresa
FAZA    = "SD"                            # faza de proiectare (SD / DTAC / PT ...)
PROIECTANT = "PROIECTANT SRL"             # completeaza cu numele firmei de proiectare

# Note standard (text propriu, completeaza notele preluate din template)
NOTA_ARMATURA = "Armatura este estimata pe baza de indici de consum (kg/mc)."
NOTA_ACOPERIS = "Nu sunt cuprinse elementele de acoperis (confectii metalice / structura de lemn)."
TEMPLATE_NOTE = ""                        # nume fisier template de unde se preiau notele (gol = fara)

def citeste_note_template(folder):
    """Preia textele notelor/observatiilor din template (daca exista)."""
    res = {"supra": [], "infra": [], "infra_hidro": None}
    path = os.path.join(folder, TEMPLATE_NOTE)
    if not os.path.exists(path):
        return res
    try:
        twb = openpyxl.load_workbook(path, data_only=True)
        def clean(v):
            return v.strip().lstrip("*").strip() if isinstance(v, str) else None
        if "SUPRA T1" in twb.sheetnames:
            ws = twb["SUPRA T1"]
            for r in (48, 49, 50):
                c = clean(ws.cell(r, 1).value)
                if c: res["supra"].append(c)
        if "INFRA" in twb.sheetnames:
            ws = twb["INFRA"]
            h = clean(ws.cell(25, 1).value)
            if h: res["infra_hidro"] = h
            for r in (26, 27, 28):
                c = clean(ws.cell(r, 1).value)
                if c: res["infra"].append(c)
        twb.close()
    except Exception as e:
        print(f"  (nu am putut citi notele din template: {e})")
    return res

# Clase beton fixe per tip element (ignora ce vine din IFC)
# None = preia din IFC
CLASE_INFRA = {
    "radier":    "C35/45 (XC2+XF3)",
    "pereti":    "C35/45 (XC2)",
    "grinzi":    "C35/45 (XC2)",
    "placi":     "C35/45 (XC2+XF3)",   # placi pastrate tot C35/45
    "atic":      "C35/45 (XC2)",
    "egalizare": "C12/15",
}
# Indici de armare infrastructura (kg/mc)
INDICI_INFRA = {
    "radier": {"_default": 135},
    "pereti": {"_default": 150},
    "grinzi": {"_default": 200},
    "placi":  {"_default": 110},
    # egalizare = beton simplu, nearmat -> fara indice
}
CLASE_SUPRA = {
    "pereti": "C30/37 (XC1)",
    "grinzi": "C30/37 (XC1)",
    "placi":  "C30/37 (XC1)",
    "atic":   "C30/37 (XC1)",
}

# Indici de armare (kg/mc) per categorie si nivel. Cheia nivelului se potriveste
# dupa eticheta fara prefixul "N - " (ex: "Parter dx"). "_default" = restul nivelurilor.
# Se scriu ca valori in coloana "indice armare"; armatura kg = beton * indice.
INDICI_SUPRA = {
    "pereti": {"Parter": 180, "Parter dx": 180, "_default": 130},  # stalpi inclusi
    "grinzi": {"_default": 135},
    "placi":  {"Etaj 1": 110, "_default": 105},                    # placa peste Parter dx (cota Etaj 1)
}
# Corpuri FARA Parter dx (ex: Nord) — nu exista "placa peste parter dx", deci placi toate 105
INDICI_SUPRA_NODX = {
    "pereti": {"Parter": 180, "_default": 130},
    "grinzi": {"_default": 135},
    "placi":  {"_default": 105},
}

# Ordinea sheet-urilor si IFC-ul asociat. EXEMPLU de configurare — inlocuieste
# numele fisierelor IFC (din folderul scriptului) cu cele ale proiectului tau.
#   "tip": "cantitativ"  -> sheet de tip lista de cantitati (incinta/consolidare)
#   "ifc": None          -> sheet gol (structura, completat ulterior)
#   "supr_inherit"       -> nivel cu placi partiale preia suprafata de la alt nivel
#   "majorari"           -> majoreaza cantitatile unui nivel cu un factor (nivel nemodelat)
SHEETS = [
    {"name": "Incinta si Consolidare teren", "tip": "cantitativ"},
    {"name": "Infrastructura",      "titlu": "INFRASTRUCTURA",                    "ifc": "infrastructura.ifc",  "clase": CLASE_INFRA, "indici": INDICI_INFRA},
    {"name": "Suprastructura T1.E", "titlu": "SUPRASTRUCTURA — Corp Est (T1.E)",  "ifc": "corp-est.ifc",        "clase": CLASE_SUPRA, "indici": INDICI_SUPRA,
     # Nivele cu suprafata redusa (placi partiale) -> preia suprafata de la nivelul indicat
     "supr_inherit": {"4 - Parter": "5 - Parter dx"}},
    {"name": "Suprastructura T1.V", "titlu": "SUPRASTRUCTURA — Corp Vest (T1.V)", "ifc": "corp-vest.ifc",       "clase": CLASE_SUPRA, "indici": INDICI_SUPRA,
     "supr_inherit": {"4 - Parter": "5 - Parter dx"}},
    {"name": "Suprastructura T1.N", "titlu": "SUPRASTRUCTURA — Corp Nord (T1.N)", "ifc": "corp-nord.ifc",       "clase": CLASE_SUPRA, "indici": INDICI_SUPRA_NODX,
     "majorari": {"Etaj 1": 1.5}},  # exemplu: Etaj 1 +50% — alocare pentru un nivel nemodelat
]

FACTOR_BETON  = 1.05
FACTOR_COFRAJ = 1.05

# ===========================================================================
# LAYOUT COLOANE
# ===========================================================================
CA = 1;  CB = 2;  CC = 3          # Nivel | Element | Clasa beton
CD = 4;  CE = 5                   # Beton mc | mc/mp
CF = 6;  CG = 7                   # Cofraj mp | mp/mp
CH = 8;  CI = 9;  CJ = 10        # Arm kg | kg/mc | kg/mp
CK = 11                           # Suprafata construita
# col 12 = separator
CM = 13; CN = 14                  # Cant. calc. beton | indice
CO = 15; CP = 16                  # Cant. calc. cofraj | indice  (coloana 15 nu mai e gol)
CR = 17                           # Indice armare

TOTAL_COLS = 17

def col(n): return get_column_letter(n)

# ===========================================================================
# STILURI
# ===========================================================================
C_DARK  = "1F4E79"
C_MID   = "2E75B6"
C_LIGHT = "BDD7EE"
C_ROW0  = "FFFFFF"
C_ROW1  = "EBF3FB"
C_TOTAL = "D6E4F0"
C_GRAND = "1F4E79"
C_AUX   = "F0F0F0"

thin  = Side(style="thin",   color="B0C4DE")
thick = Side(style="medium", color=C_DARK)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_SEP = Border(left=thin, right=thick, top=thin, bottom=thin)  # separator dupa K

def sc(cell, fill=None, font=None, align="center", num_fmt=None, sep=False, no_border=False):
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = font or Font(size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if not no_border:
        cell.border = BORDER_SEP if sep else BORDER
    if num_fmt:
        cell.number_format = num_fmt

def hdr_font(size=10, color="FFFFFF"):
    return Font(bold=True, color=color, size=size)

def bold_font(size=10, color="000000"):
    return Font(bold=True, size=size, color=color)

# ===========================================================================
# IFC — procesare
# ===========================================================================
GEOM_SETTINGS = None
_type_cache   = None

def init_geom():
    global GEOM_SETTINGS
    s = ifcopenshell.geom.settings()
    s.set(s.USE_WORLD_COORDS, True)
    GEOM_SETTINGS = s

def build_type_cache(ifc):
    global _type_cache
    _type_cache = {}
    for rel in ifc.by_type("IfcRelDefinesByType"):
        for el in rel.RelatedObjects:
            _type_cache[el.id()] = rel.RelatingType.Name or ""

def get_type_name(el):
    return _type_cache.get(el.id(), "") if _type_cache else ""

def extract_clasa_beton(type_name):
    m = re.search(r'C\s*\d+/\d+', type_name or "")
    return m.group(0) if m else "-"

def nivel_label(name):
    # "5 - Parter dx" -> "Parter dx"
    return re.sub(r'^\s*\d+\s*-\s*', '', name or '').strip()

def get_indice(indici, cat, level_name):
    """Indicele de armare (kg/mc) pt o categorie pe un nivel, sau None."""
    if not indici:
        return None
    rules = indici.get(cat, {})
    lbl = nivel_label(level_name)
    if lbl in rules:
        return rules[lbl]
    return rules.get("_default")

CAT_LABEL = {"pereti":"Pereti BA","grinzi":"Grinzi BA",
             "placi":"Placi, scari, reborduri","atic":"Atic BA",
             "egalizare":"Beton simplu egalizare","radier":"Radier"}
CAT_ORDER = ["egalizare","radier","pereti","grinzi","placi","atic"]

def is_stalp_beton(tn):
    # stalp de beton: "Ortbeton" / contine clasa C##/## / "BA"; otelul (RO/EN tubular) = nu
    t = (tn or "").lower()
    if "ortbeton" in t or "beton" in t or " ba" in t.lower():
        return True
    return bool(re.search(r'C\s*\d+/\d+', tn or ""))

def categorize(el):
    cls = el.is_a(); tn = get_type_name(el)
    if cls in ("IfcWall","IfcWallStandardCase"):
        return "placi" if "Atic" in (tn or "") else "pereti"  # atic -> adaugat la placi
    if cls == "IfcColumn":
        return "pereti" if is_stalp_beton(tn) else None  # stalpi beton -> la pereti; otel -> exclus
    if cls == "IfcBeam": return "grinzi"
    if cls == "IfcSlab":
        t = (tn or "").lower()
        if "egalizare" in t: return "egalizare"
        if "radier"    in t: return "radier"
        return "placi"
    return None

def compute_geometry(el):
    try: shape = ifcopenshell.geom.create_shape(GEOM_SETTINGS, el)
    except: return None, None, None, None, None
    verts = shape.geometry.verts; faces = shape.geometry.faces
    coords = [(verts[i*3], verts[i*3+1], verts[i*3+2]) for i in range(len(verts)//3)]
    vol = area_tot = area_h = 0.0
    for i in range(len(faces)//3):
        a,b,c = coords[faces[i*3]],coords[faces[i*3+1]],coords[faces[i*3+2]]
        ab=(b[0]-a[0],b[1]-a[1],b[2]-a[2]); ac=(c[0]-a[0],c[1]-a[1],c[2]-a[2])
        nx=ab[1]*ac[2]-ab[2]*ac[1]; ny=ab[2]*ac[0]-ab[0]*ac[2]; nz=ab[0]*ac[1]-ab[1]*ac[0]
        mag=math.sqrt(nx*nx+ny*ny+nz*nz); area_tri=0.5*mag; area_tot+=area_tri
        if mag>0 and abs(nz)/mag>0.9: area_h+=area_tri
        vol+=(a[0]*nx+a[1]*ny+a[2]*nz)/6.0
    zs = [c[2] for c in coords]
    return abs(vol), abs(area_tot), area_h, min(zs), max(zs)

def cofraj_val(cls, area_tot, area_h):
    if cls in ("IfcWall","IfcWallStandardCase","IfcBeam","IfcColumn"): return area_tot - area_h
    if cls == "IfcSlab": return area_h / 2.0
    return area_tot

def build_storey_map(ifc):
    m = {}
    for rel in ifc.by_type("IfcRelContainedInSpatialStructure"):
        if rel.RelatingStructure.is_a("IfcBuildingStorey"):
            sn = rel.RelatingStructure.Name or rel.RelatingStructure.GlobalId
            for el in rel.RelatedElements: m[el.id()] = sn
    return m

def process_ifc(ifc_path):
    print(f"  Procesez: {os.path.basename(ifc_path)}")
    ifc = ifcopenshell.open(ifc_path)
    init_geom(); build_type_cache(ifc)
    storeys = sorted(ifc.by_type("IfcBuildingStorey"),
                     key=lambda s: float(s.Elevation) if s.Elevation else 0.0)
    storey_order = [s.Name or s.GlobalId for s in storeys]
    storey_map   = build_storey_map(ifc)

    # Cote nivele (pt asignarea peretilor dupa placa sustinuta)
    elevs = [(s.Name or s.GlobalId, float(s.Elevation) if s.Elevation else 0.0) for s in storeys]
    def nivel_dupa_cota(z):
        # nivelul a carui cota e cea mai apropiata de z
        return min(elevs, key=lambda ne: abs(ne[1] - z))[0]

    rezultate    = defaultdict(lambda: defaultdict(lambda: {"clasa":"-","vol":0.0,"cof":0.0,"nr":0}))
    supr_nivel   = defaultdict(float)
    elemente     = (ifc.by_type("IfcWall") + ifc.by_type("IfcColumn")
                    + ifc.by_type("IfcBeam") + ifc.by_type("IfcSlab"))
    total, skip  = len(elemente), 0
    reassigned   = 0
    for idx, el in enumerate(elemente):
        if idx % 200 == 0: print(f"    {idx}/{total}...", end="\r")
        cat = categorize(el)
        if not cat: continue
        vol, area_tot, area_h, minz, maxz = compute_geometry(el)
        if vol is None: skip += 1; continue

        # Elementele verticale (pereti, atic, stalpi beton) se asigneaza nivelului a carui
        # PLACA o sustin (cota varf), nu nivelului pe care stau (cota baza, cum face Revit).
        # Grinzile si placile raman la nivelul lor Revit (sunt corecte).
        is_vertical = el.is_a() in ("IfcWall", "IfcWallStandardCase", "IfcColumn")
        if is_vertical:
            sname = nivel_dupa_cota(maxz)
            if sname != storey_map.get(el.id()): reassigned += 1
        else:
            sname = storey_map.get(el.id(), "Fara nivel")

        cof = cofraj_val(el.is_a(), area_tot, area_h)
        if cat == "egalizare":
            cof = 0.0                       # beton simplu turnat pe teren — fara cofraj
        elif cat == "radier":
            cof = area_tot - area_h         # doar cofrajul lateral (cant), intradosul e pe egalizare

        d = rezultate[sname][cat]
        d["clasa"] = extract_clasa_beton(get_type_name(el))
        d["vol"]  += vol; d["cof"] += cof; d["nr"] += 1
        # suprafata construita = amprenta placilor reale (NU egalizarea, aceeasi amprenta)
        if el.is_a() == "IfcSlab" and cat != "egalizare":
            supr_nivel[sname] += area_h / 2.0
    print(f"    Gata: {total-skip}/{total} elemente | pereti reasignati dupa varf: {reassigned}")
    return rezultate, storey_order, supr_nivel

# ===========================================================================
# EXCEL — scriere un sheet
# ===========================================================================
HEADER_ROWS = 6   # randuri de antet inainte de date

def write_sheet(ws, titlu, rezultate, storey_order, supr_nivel, clase=None, indici=None,
                majorari=None, notes=None):
    """Scrie un sheet complet. rezultate=None => sheet gol cu structura."""

    all_levels = [l for l in storey_order if l in (rezultate or {})]
    if rezultate:
        for l in rezultate:
            if l not in all_levels: all_levels.append(l)

    # ----------------------------------------------------------------
    # ANTET (randuri 1-6) — zona printabila
    # ----------------------------------------------------------------
    for i in range(1, HEADER_ROWS+1):
        ws.row_dimensions[i].height = 18

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 36
    ws.row_dimensions[6].height = 16

    # R1: Proiectant + faza (stanga) | Titlu sheet (centru) | Proiect (dreapta)
    # Parte stanga: A-C
    ws.merge_cells("A1:C1")
    c = ws.cell(row=1, column=CA, value=PROIECTANT)
    sc(c, fill=C_DARK, font=hdr_font(11), align="left"); c.border = Border()

    # Centru: D-H titlu sheet
    ws.merge_cells("D1:H1")
    c = ws.cell(row=1, column=CD, value=titlu)
    sc(c, fill=C_DARK, font=hdr_font(12), align="center"); c.border = Border()

    # Dreapta: I-K proiect
    ws.merge_cells("I1:K1")
    c = ws.cell(row=1, column=CI, value=f"FAZA: {FAZA}")
    sc(c, fill=C_DARK, font=hdr_font(10), align="right"); c.border = Border()

    # R2: Denumire proiect
    ws.merge_cells("A2:K2")
    c = ws.cell(row=2, column=CA, value=PROIECT)
    sc(c, fill=C_MID, font=hdr_font(10), align="center"); c.border = Border()

    # R3: Adresa (inlocuim "A. Lucrari structura" cu split pe 2 randuri)
    ws.merge_cells("A3:E3")
    c = ws.cell(row=3, column=CA, value=ADRESA)
    sc(c, fill=C_DARK, font=Font(italic=True, size=9, color="CCDDEE"), align="left"); c.border = Border()
    ws.merge_cells("F3:K3")
    c = ws.cell(row=3, column=CF, value="A. Lucrari structura")
    sc(c, fill=C_DARK, font=hdr_font(10), align="right"); c.border = Border()

    # R4: grupuri coloane
    ws.merge_cells("D4:E4"); ws.cell(row=4,column=CD).value="Beton"
    ws.merge_cells("F4:G4"); ws.cell(row=4,column=CF).value="Cofraj"
    ws.merge_cells("H4:J4"); ws.cell(row=4,column=CH).value="Armatura"
    ws.cell(row=4,column=CK).value="Suprafata\nconstruita\n/ nivel"

    grp4 = [(CA,CB,"",C_DARK),(CC,CC,"",C_DARK),(CD,CE,"Beton",C_MID),
            (CF,CG,"Cofraj",C_MID),(CH,CJ,"Armatura",C_MID),(CK,CK,"Suprafata\nconstruita\n/ nivel",C_MID)]
    for c1,c2,txt,fill in grp4:
        if c1!=c2: ws.merge_cells(start_row=4,start_column=c1,end_row=4,end_column=c2)
        cell=ws.cell(row=4,column=c1,value=txt)
        sc(cell,fill=fill,font=hdr_font(10))

    # R5: sub-anteturi
    hdrs5={CA:"Nivel",CB:"Element",CC:"Clasa beton",
           CD:"mc",CE:"mc/mp",CF:"mp",CG:"mp/mp",
           CH:"kg",CI:"kg/mc",CJ:"kg/mp",CK:"mp"}
    for ci,txt in hdrs5.items():
        cell=ws.cell(row=5,column=ci,value=txt)
        sc(cell,fill=C_MID,font=hdr_font(9))

    # R6: linie fina (goala)
    for ci in range(1,CK+1):
        cell=ws.cell(row=6,column=ci)
        sc(cell,fill=C_DARK,font=Font(size=2))
    ws.row_dimensions[6].height = 4

    # Coloane auxiliare — antet compact
    aux_hdrs = {CM:"mc (IFC)",CN:"coef.",CO:"mp (IFC)",CP:"coef.",CR:"kg/mc"}
    ws.merge_cells(f"M4:{col(CR)}4")
    c_aux_hdr=ws.cell(row=4,column=CM,value="Zone de calcul auxiliare  (nu se printeaza)")
    sc(c_aux_hdr,fill="CCCCCC",font=Font(italic=True,size=8,color="555555"),align="left",no_border=True)
    for ci,txt in aux_hdrs.items():
        cell=ws.cell(row=5,column=ci,value=txt)
        sc(cell,fill=C_AUX,font=Font(bold=True,size=8,color="555555"))

    DATA_START = 7

    if not rezultate:
        # Sheet gol — doar structura
        ws.print_area = f"A1:K{DATA_START+2}"
        return None

    # ----------------------------------------------------------------
    # DATE
    # ----------------------------------------------------------------
    row = DATA_START

    for level in all_levels:
        cat_data = rezultate.get(level, {})
        cats_present = [c for c in CAT_ORDER if cat_data.get(c,{}).get("nr",0)>0]
        if not cats_present: continue

        supr = supr_nivel.get(level, 0)
        level_start = row
        level_end   = row + len(cats_present) - 1
        k_ref = f"${col(CK)}${level_start}"
        # factor de majorare nivel (ex: Nord Etaj 1 x1.5 pt mansarda nemodelata)
        factor = (majorari or {}).get(nivel_label(level), 1.0)
        fx = "" if factor == 1.0 else f"*{factor}"

        for ci, cat_key in enumerate(cats_present):
            d     = cat_data[cat_key]
            label = CAT_LABEL.get(cat_key, cat_key)
            vol   = round(d["vol"], 3)
            cof_v = round(d["cof"], 2)
            r     = row
            fill  = C_ROW0 if ci%2==0 else C_ROW1

            # Valori brute in coloane auxiliare
            ws.cell(row=r, column=CM, value=vol)
            ws.cell(row=r, column=CN, value=FACTOR_BETON)
            ws.cell(row=r, column=CO, value=cof_v)
            ws.cell(row=r, column=CP, value=FACTOR_COFRAJ)
            idx = get_indice(indici, cat_key, level)        # indice armare kg/mc
            ws.cell(row=r, column=CR, value=idx if idx is not None else "")

            # Formule coloane principale
            ws.cell(row=r, column=CB,  value=label)
            # Placi: +3mc/nivel pt scari nemodulate. fx = majorare nivel (vizibila in formula).
            base_b = f"{col(CM)}{r}*{col(CN)}{r}" + ("+3" if cat_key=="placi" else "")
            base_c = f"{col(CO)}{r}*{col(CP)}{r}" + ("+20" if cat_key=="placi" else "")
            if fx:
                ws.cell(row=r, column=CD, value=f"=({base_b}){fx}")
                ws.cell(row=r, column=CF, value=f"=({base_c}){fx}")
            else:
                ws.cell(row=r, column=CD, value=f"={base_b}")
                ws.cell(row=r, column=CF, value=f"={base_c}")
            ws.cell(row=r, column=CE,  value=f"=IFERROR({col(CD)}{r}/{k_ref},\"-\")")
            ws.cell(row=r, column=CG,  value=f"=IFERROR({col(CF)}{r}/{k_ref},\"-\")")
            ws.cell(row=r, column=CH,  value=f"=IF({col(CR)}{r}>0,{col(CD)}{r}*{col(CR)}{r},\"\")")
            ws.cell(row=r, column=CI,  value=f"=IFERROR({col(CH)}{r}/{col(CD)}{r},\"\")")

            # Stiluri celule principale (A-K) — stil UNITAR, fara diferenta armat
            for col_i in range(CA, CK+1):
                cell = ws.cell(row=r, column=col_i)
                is_sep = (col_i == CK)
                if col_i == CA:
                    sc(cell, fill=C_LIGHT, font=bold_font(color=C_DARK), sep=False)
                else:
                    sc(cell, fill=fill, font=Font(size=10), sep=is_sep,
                       num_fmt="#,##0.00" if col_i in (CD,CF) else
                               "0.00"     if col_i in (CE,CG) else
                               "#,##0"    if col_i == CH else
                               "0"        if col_i in (CI,CJ) else None)

            # Stiluri auxiliare (M-R)
            for col_i in (CM,CN,CO,CP,CR):
                cell=ws.cell(row=r,column=col_i)
                sc(cell,fill=C_AUX,font=Font(size=9,color="555555"),
                   num_fmt="#,##0.00" if col_i in (CM,CO) else "0.00")

            row += 1

        # Merge A (Nivel)
        if len(cats_present)>1: ws.merge_cells(start_row=level_start,start_column=CA,end_row=level_end,end_column=CA)
        cv=ws.cell(row=level_start,column=CA); cv.value=level
        cv.fill=PatternFill("solid",fgColor=C_LIGHT)
        cv.font=bold_font(color=C_DARK)
        cv.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cv.border=BORDER

        # Clasa beton (C): clasa fixa per categorie daca e configurata, altfel din IFC.
        # Merge pe nivel doar daca toate categoriile au aceeasi clasa; altfel pe rand
        # (ex: infrastructura — egalizare C12/15 difera de radier C35/45).
        def class_for(ck):
            return (clase.get(ck) if clase else None) or cat_data[ck]["clasa"]
        row_classes = [class_for(ck) for ck in cats_present]
        if len(set(row_classes)) == 1:
            if len(cats_present) > 1:
                ws.merge_cells(start_row=level_start,start_column=CC,end_row=level_end,end_column=CC)
            cc_cell=ws.cell(row=level_start,column=CC); cc_cell.value=row_classes[0]
            cc_cell.fill=PatternFill("solid",fgColor=C_ROW0); cc_cell.font=Font(size=10)
            cc_cell.alignment=Alignment(horizontal="center",vertical="center"); cc_cell.border=BORDER
        else:
            for i, cv2 in enumerate(row_classes):
                cc_cell=ws.cell(row=level_start+i,column=CC); cc_cell.value=cv2
                cc_cell.fill=PatternFill("solid",fgColor=C_ROW0); cc_cell.font=Font(size=10)
                cc_cell.alignment=Alignment(horizontal="center",vertical="center"); cc_cell.border=BORDER

        # Merge K (Suprafata nivel) — cu separator dreapta
        if len(cats_present)>1: ws.merge_cells(start_row=level_start,start_column=CK,end_row=level_end,end_column=CK)
        ck_cell=ws.cell(row=level_start,column=CK)
        ck_cell.value=round(supr,0) if supr>0 else "-"
        ck_cell.fill=PatternFill("solid",fgColor=C_LIGHT); ck_cell.font=bold_font(color=C_DARK)
        ck_cell.alignment=Alignment(horizontal="center",vertical="center")
        ck_cell.border=BORDER_SEP

        # Merge J (kg/mp nivel) — formula SUM H / K
        h_range=f"{col(CH)}{level_start}:{col(CH)}{level_end}"
        if len(cats_present)>1: ws.merge_cells(start_row=level_start,start_column=CJ,end_row=level_end,end_column=CJ)
        cj_cell=ws.cell(row=level_start,column=CJ)
        cj_cell.value=f"=IFERROR(SUM({h_range})/{k_ref},\"\")"
        sc(cj_cell,fill=C_LIGHT if supr>0 else C_ROW0,font=Font(size=10),num_fmt="0")

        row += 1  # spatiu intre nivele

    # ----------------------------------------------------------------
    # TOTAL GENERAL
    # ----------------------------------------------------------------
    last_data = row - 2
    tot_row   = row

    ws.merge_cells(f"A{tot_row}:C{tot_row}")
    ct=ws.cell(row=tot_row,column=CA,value="Total [fara beton simplu]")
    sc(ct,fill=C_GRAND,font=hdr_font(11))
    ws.row_dimensions[tot_row].height=22

    d_rng=f"{col(CD)}{DATA_START}:{col(CD)}{last_data}"
    f_rng=f"{col(CF)}{DATA_START}:{col(CF)}{last_data}"
    h_rng=f"{col(CH)}{DATA_START}:{col(CH)}{last_data}"
    k_rng=f"{col(CK)}{DATA_START}:{col(CK)}{last_data}"
    m_rng=f"{col(CM)}{DATA_START}:{col(CM)}{last_data}"
    o_rng=f"{col(CO)}{DATA_START}:{col(CO)}{last_data}"
    t=tot_row

    totale={
        CD:f"=SUM({d_rng})",
        CE:f"=IFERROR({col(CD)}{t}/{col(CK)}{t},\"-\")",
        CF:f"=SUM({f_rng})",
        CG:f"=IFERROR({col(CF)}{t}/{col(CK)}{t},\"-\")",
        CH:f"=SUM({h_rng})",
        CI:f"=IFERROR({col(CH)}{t}/{col(CD)}{t},\"\")",
        CJ:f"=IFERROR({col(CH)}{t}/{col(CK)}{t},\"\")",
        CK:f"=SUM({k_rng})",
        CM:f"=SUM({m_rng})",
        CN:f"=IFERROR({col(CD)}{t}/{col(CM)}{t},\"\")",
        CO:f"=SUM({o_rng})",
        CP:f"=IFERROR({col(CF)}{t}/{col(CO)}{t},\"\")",
        CR:"",
    }
    for ci,formula in totale.items():
        cell=ws.cell(row=t,column=ci,value=formula)
        is_sep=(ci==CK)
        is_aux=(ci in (CM,CN,CO,CP,CR))
        if is_aux:
            sc(cell,fill=C_AUX,font=Font(bold=True,size=9,color="555555"),
               num_fmt="#,##0.00" if ci in (CM,CO) else "0.00")
        else:
            sc(cell,fill=C_GRAND,font=hdr_font(11),sep=is_sep,
               num_fmt="#,##0.00" if ci in (CD,CF) else
                       "0.00"     if ci in (CE,CG) else
                       "#,##0"    if ci == CH else
                       "0"        if ci in (CI,CJ) else None)

    # Rand unitati
    unit_row=tot_row+1
    ws.merge_cells(f"A{unit_row}:C{unit_row}")
    for ci,txt in {CD:"mc",CE:"mc/mp",CF:"mp",CG:"mp/mp",CH:"kg",CI:"kg/mc",CJ:"kg/mp",CK:"mp"}.items():
        cell=ws.cell(row=unit_row,column=ci,value=txt)
        sc(cell,fill="F2F2F2",font=Font(italic=True,size=9,color="888888"))

    # Note (sub tabel) — inaltime dinamica pt note lungi (ex: hidroizolatie infra)
    last_row = unit_row
    for nt in (notes or []):
        last_row += 1
        ws.merge_cells(start_row=last_row,start_column=CA,end_row=last_row,end_column=CK)
        cell=ws.cell(row=last_row,column=CA,value="* "+nt)
        cell.fill=PatternFill("solid",fgColor="FFF8E1")
        cell.font=Font(italic=True,size=9,color="7A5C00")
        cell.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
        cell.border=BORDER
        txt="* "+nt
        nlines=txt.count("\n")+max(1, math.ceil(len(txt)/145))
        ws.row_dimensions[last_row].height=max(18, nlines*12)

    # ----------------------------------------------------------------
    # PRINT SETUP
    # ----------------------------------------------------------------
    ws.print_area             = f"A1:K{last_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 99
    ws.page_margins.left      = 0.5
    ws.page_margins.right     = 0.5
    ws.page_margins.top       = 0.6
    ws.page_margins.bottom    = 0.6

    # Freeze sub antet
    ws.freeze_panes = f"D{DATA_START}"

    return tot_row   # randul "Total" — folosit de centralizator

# ===========================================================================
# LATIMI COLOANE (comune pentru toate sheet-urile)
# ===========================================================================
COL_WIDTHS = {
    CA:16, CB:22, CC:14,
    CD:11, CE:10, CF:11, CG:10,
    CH:11, CI:10, CJ:10, CK:14,
    12:3,                         # separator vizual
    CM:11, CN:7, CO:11, CP:7,
    16:3,                         # separator vizual
    CR:9,
}

def apply_col_widths(ws):
    for ci,w in COL_WIDTHS.items():
        ws.column_dimensions[col(ci)].width=w
    # Ascunde col 12 si 16 (separatoare)
    ws.column_dimensions[col(12)].hidden=True
    ws.column_dimensions[col(16)].hidden=True

# ===========================================================================
# CENTRALIZATOR — totaluri pe corpuri + cost estimativ euro/mp
# ===========================================================================
# Preturi unitare ESTIMATIVE (editabile in Excel). Inlocuieste cu preturile tale.
PRET_BETON   = 0.0     # €/mc beton armat — completeaza cu pretul proiectului (material + manopera)
PRET_COFRAJ  = 0.0     # €/mp cofraj — completeaza
PRET_ARM     = 0.0     # €/kg armatura — completeaza

def write_centralizator(ws, corpuri):
    """corpuri = list de (eticheta, sheet_name, tot_row)."""
    # Col: A Corp | B Beton mc | C Cofraj mp | D Armatura kg | E Supr mp |
    #      F Cost beton | G Cost cofraj | H Cost armatura | I Cost total | J €/mp
    W = {1:26,2:12,3:12,4:13,5:12,6:14,7:14,8:14,9:15,10:11}
    for c,w in W.items(): ws.column_dimensions[col(c)].width=w

    ws.merge_cells("A1:J1")
    c=ws["A1"]; c.value="CENTRALIZATOR CANTITATI SI COST ESTIMATIV STRUCTURA"
    sc(c, fill=C_DARK, font=hdr_font(13)); ws.row_dimensions[1].height=26
    ws.merge_cells("A2:J2")
    c=ws["A2"]; c.value=PROIECT
    sc(c, fill=C_MID, font=hdr_font(10))

    # Preturi unitare (editabile)
    preturi=[("Pret unitar beton",   "€/mc", PRET_BETON,  3),
             ("Pret unitar cofraj",  "€/mp", PRET_COFRAJ, 4),
             ("Pret unitar armatura","€/kg", PRET_ARM,    5)]
    ws.merge_cells("A3:A5")
    cc=ws["A3"]; cc.value="Preturi unitare\n(editabile)"
    sc(cc, fill=C_LIGHT, font=bold_font(color=C_DARK))
    for lbl,um,val,r in preturi:
        ws.cell(row=r,column=2,value=lbl);
        sc(ws.cell(row=r,column=2), fill=C_ROW1, align="left")
        cv=ws.cell(row=r,column=3,value=val)
        sc(cv, fill="FFF2CC", font=bold_font()); cv.number_format='#,##0.00'
        ws.cell(row=r,column=4,value=um)
        sc(ws.cell(row=r,column=4), fill=C_ROW1, align="left", font=Font(italic=True,size=9,color="888888"))
        for cx in (5,6,7,8,9,10):
            sc(ws.cell(row=r,column=cx), fill=C_ROW0)
    P_BET=f"$C$3"; P_COF=f"$C$4"; P_ARM=f"$C$5"

    # Antet tabel (rand 7)
    hdr={1:"Corp / Obiect",2:"Beton",3:"Cofraj",4:"Armatura",5:"Supr.\nconstr.",
         6:"Cost beton",7:"Cost cofraj",8:"Cost arm.",9:"Cost total",10:"€/mp"}
    um ={1:"",2:"mc",3:"mp",4:"kg",5:"mp",6:"€",7:"€",8:"€",9:"€",10:"€/mp"}
    ws.row_dimensions[7].height=24
    for cx,t in hdr.items(): sc(ws.cell(row=7,column=cx,value=t), fill=C_MID, font=hdr_font(10))
    for cx,t in um.items():  sc(ws.cell(row=8,column=cx,value=t), fill="F2F2F2", font=Font(italic=True,size=9,color="888888"))

    # Randuri corpuri (rand 9+)
    r=9; first=r
    for eticheta, sheet_name, tot in corpuri:
        q = f"'{sheet_name}'!"
        ws.cell(row=r,column=1,value=eticheta)
        ws.cell(row=r,column=2,value=f"={q}D{tot}")   # beton mc
        ws.cell(row=r,column=3,value=f"={q}F{tot}")   # cofraj mp
        ws.cell(row=r,column=4,value=f"={q}H{tot}")   # armatura kg
        ws.cell(row=r,column=5,value=f"={q}K{tot}")   # suprafata mp
        ws.cell(row=r,column=6,value=f"=B{r}*{P_BET}")
        ws.cell(row=r,column=7,value=f"=C{r}*{P_COF}")
        ws.cell(row=r,column=8,value=f"=D{r}*{P_ARM}")
        ws.cell(row=r,column=9,value=f"=F{r}+G{r}+H{r}")
        ws.cell(row=r,column=10,value=f"=IFERROR(I{r}/E{r},\"-\")")
        fill = C_ROW0 if (r-first)%2==0 else C_ROW1
        for cx in range(1,11):
            cell=ws.cell(row=r,column=cx)
            sc(cell, fill=fill, align="left" if cx==1 else "center")
            if cx in (2,3,4,5,6,7,8,9): cell.number_format='#,##0'
            if cx==10: cell.number_format='#,##0'
        r+=1

    # Total
    last=r-1
    ws.cell(row=r,column=1,value="TOTAL GENERAL")
    for cx,letter in [(2,'B'),(3,'C'),(4,'D'),(5,'E'),(6,'F'),(7,'G'),(8,'H'),(9,'I')]:
        ws.cell(row=r,column=cx,value=f"=SUM({letter}{first}:{letter}{last})")
    ws.cell(row=r,column=10,value=f"=IFERROR(I{r}/E{r},\"-\")")
    for cx in range(1,11):
        cell=ws.cell(row=r,column=cx)
        sc(cell, fill=C_GRAND, font=hdr_font(11), align="left" if cx==1 else "center")
        if cx in range(2,10): cell.number_format='#,##0'
        if cx==10: cell.number_format='#,##0'
    ws.row_dimensions[r].height=22
    total_row=r

    # Nota
    r+=2
    ws.merge_cells(f"A{r}:J{r}")
    nota=ws.cell(row=r,column=1,value="Preturile unitare sunt estimative si editabile (celulele galbene C3:C5). "
                "Cantitatile sunt preluate automat din sheet-urile fiecarui corp. "
                "Armatura rezulta din indicii de armare; €/mp = cost total / suprafata construita desfasurata. "
                "NU sunt cuprinse elementele de acoperis (confectii metalice / structura de lemn). "
                "Corp Nord: Etajul 1 este majorat cu ~50% ca alocare pentru mansarda nemodelata.")
    sc(nota, fill=C_ROW0, align="left", font=Font(italic=True,size=9,color="666666"))
    ws.row_dimensions[r].height=42

    # Print
    ws.print_area=f"A1:J{total_row}"
    ws.page_setup.orientation="landscape"
    ws.page_setup.fitToPage=True; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=99
    return total_row

# ===========================================================================
# CANTITATIV DE LUCRARI — INCINTA SI CONSOLIDARE TEREN
# ===========================================================================
# Articol = (denumire, UM, formula(r)->str, mod_de_calcul). Cantitatea e formula Excel.
# Taluz: volum = lungime * (proiectie_orizontala * inaltime / 2)  [sectiune triunghiulara]
# VALORILE de mai jos sunt EXEMPLU — inlocuieste dimensiunile/numerele cu cele ale proiectului.
# Indicii de armare din formule (160, 120, 50 kg/...) sunt valori uzuale, ajusteaza-le.
ARTICOLE_INCINTA = [
    ("Sapatura generala mecanizata - ETAPA 1", "mc",
        lambda r: "=5000*3+300*2*3/2", "Aria x adancime + taluz: perimetru x (proiectie x inaltime / 2)"),
    ("Sapatura generala mecanizata - ETAPA 2", "mc",
        lambda r: "=4000*5+150*4*5/2", "Aria x adancime + taluz: lungime x (proiectie x inaltime / 2)"),
    ("Umplutura de pamant compactat (din taluzuri)", "mc",
        lambda r: f"=(D{r-2}-5000*3)+(D{r-1}-4000*5)", "Volum taluzuri ETAPA 1 + ETAPA 2 (reutilizate ca umplutura)"),
    ("Foraj piloti Ø60cm - tehnologie CFA/bentonita", "ml",
        lambda r: "=100*12", "nr. piloti x lungime"),
    ("Beton armat C30/37 in piloti Ø60cm", "mc",
        lambda r: "=100*PI()*0.3^2*12", "nr. x π x raza² x lungime"),
    ("Armatura piloti Ø60 BST 500C clasa de ductilitate C", "kg",
        lambda r: f"=D{r-1}*160", "Beton piloti x indice (kg/mc)"),
    ("Spargere cap piloti beton contaminat (~50cm)", "mc",
        lambda r: "=100*PI()*0.3^2*0.5", "nr. x π x raza² x adancime spargere"),
    ("Beton armat C30/37 in grinzi coronament", "mc",
        lambda r: "=80*0.6*0.8", "lungime x latime x inaltime sectiune"),
    ("Armatura grinzi coronament BST 500C clasa de ductilitate C", "kg",
        lambda r: f"=D{r-1}*120", "Beton grinzi x indice (kg/mc)"),
]
# Consolidare teren — incluziuni rigide (EXEMPLU: nr. buc, Ø, L, indice kg/buc)
ARTICOLE_CONSOLIDARE = [
    ("Executare incluziuni rigide Ø400mm, L=6.00m, tehnologie FDP", "ml",
        lambda r: "=500*6", "nr. buc x lungime"),
    ("Beton C30/37 in incluziuni rigide (turnat cu pompa)", "mc",
        lambda r: "=500*PI()*0.2^2*6", "nr. x π x raza² x lungime"),
    ("Armatura incluziuni rigide BST 500C - piloti slab armati", "kg",
        lambda r: "=500*50", "nr. buc x indice (kg/buc)"),
]
NOTE_CANTITATIV = [
    "Valorile reprezinta cantitati teoretice si nu cuprind pierderile tehnologice aferente punerii in opera.",
    "Armatura (piloti, grinzi de coronament, incluziuni rigide) este estimata pe baza de indici de consum.",
    "Torcretul (min. 5cm) pe pilotii de incinta se recomanda ca strat suport pentru hidroizolatie.",
    "Spraiturile (otel laminat S355) nu sunt incluse.",
    "Sunt necesare puturi de epuisment, care se vor dimensiona pe baza unui proiect de epuismente.",
    "In functie de concluziile studiului hidrogeologic si ale proiectului de epuismente se va stabili "
    "daca baza taluzului va fi torcretata si tintuita.",
]

# Layout cantitativ: A Nr | B Denumire | C UM | D Cantitate | E Mod de calcul
QN, QD, QU, QQ, QC = 1,2,3,4,5

def write_cantitativ(ws):
    for c,w in {QN:6,QD:64,QU:7,QQ:14,QC:44}.items():
        ws.column_dimensions[col(c)].width=w

    # Zona printabila = A:D. Coloana E (Mod de calcul) ramane pe ecran, dar in afara printului.
    ws.merge_cells("A1:C1"); c=ws.cell(1,QN,value=PROIECTANT)
    sc(c, fill=C_DARK, font=hdr_font(11), align="left"); c.border=Border()
    c=ws.cell(1,QQ,value=f"FAZA: {FAZA}")
    sc(c, fill=C_DARK, font=hdr_font(10), align="right"); c.border=Border()
    ws.merge_cells("A2:D2"); c=ws.cell(2,QN,value=PROIECT)
    sc(c, fill=C_MID, font=hdr_font(10)); c.border=Border()
    ws.merge_cells("A3:D3"); c=ws.cell(3,QN,value="CANTITATIV DE LUCRARI — INCINTA SI CONSOLIDARE TEREN")
    sc(c, fill=C_DARK, font=hdr_font(12)); ws.row_dimensions[3].height=24

    hd={QN:"Nr.",QD:"Denumire element / articol",QU:"UM",QQ:"Cantitate",QC:"Mod de calcul (uz intern, nu se printeaza)"}
    ws.row_dimensions[4].height=20
    for c,t in hd.items():
        fill = "D9D9D9" if c==QC else C_MID
        sc(ws.cell(4,c,value=t), fill=fill, font=Font(bold=True,size=9,color="555555") if c==QC else hdr_font(10),
           align="left" if c in (QD,QC) else "center")

    r=5
    def sectiune(titlu, articole):
        nonlocal r
        ws.merge_cells(start_row=r,start_column=QN,end_row=r,end_column=QQ)
        sc(ws.cell(r,QN,value=titlu), fill=C_LIGHT, font=bold_font(color=C_DARK), align="left")
        r+=1
        for i,(den,um,fml,calc) in enumerate(articole,1):
            fill = C_ROW0 if i%2 else C_ROW1
            ws.cell(r,QN,value=i)
            ws.cell(r,QD,value=den)
            ws.cell(r,QU,value=um)
            ws.cell(r,QQ,value=fml(r))
            ws.cell(r,QC,value=calc)
            for c in range(QN,QC+1):
                cell=ws.cell(r,c)
                sc(cell, fill=fill, align="left" if c in (QD,QC) else "center")
                if c==QC: cell.font=Font(size=9,color="555555")
                if c==QQ: cell.number_format='#,##0.##'
            r+=1

    sectiune("A. INCINTA (sprijiniri excavatie)", ARTICOLE_INCINTA)
    sectiune("B. CONSOLIDARE TEREN (incluziuni rigide)", ARTICOLE_CONSOLIDARE)

    # Note (in zona printabila A:D)
    for nt in NOTE_CANTITATIV:
        ws.merge_cells(start_row=r,start_column=QN,end_row=r,end_column=QQ)
        cell=ws.cell(r,QN,value="* "+nt)
        cell.fill=PatternFill("solid",fgColor="FFF8E1"); cell.font=Font(italic=True,size=9,color="7A5C00")
        cell.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True); cell.border=BORDER
        txt="* "+nt; ws.row_dimensions[r].height=max(16, (txt.count("\n")+max(1,math.ceil(len(txt)/95)))*12)
        r+=1

    ws.print_area=f"A1:D{r-1}"
    ws.page_setup.orientation="landscape"
    ws.page_setup.fitToPage=True; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
    ws.freeze_panes="A5"

# ===========================================================================
# COPIE CLIENT — doar partea printabila, formule ascunse, fara centralizator
# ===========================================================================
def make_client_copy(master_path, client_path):
    wb = openpyxl.load_workbook(master_path)
    for ws in wb.worksheets:
        if ws.title == "Centralizator":
            ws.sheet_state = "hidden"          # ascunde sheetul cu costuri
            continue
        # ascunde coloanele de calcul (in afara zonei printabile)
        hide_from = 5 if ws.title == "Incinta si Consolidare teren" else 12  # E+ / L+
        for ci in range(hide_from, 28):
            ws.column_dimensions[col(ci)].hidden = True
        # ascunde formulele (nu se vad in bara de formule cand foaia e protejata)
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    c.protection = Protection(locked=True, hidden=True)
        ws.protection.sheet = True             # fara parola (se poate adauga ulterior)
    try:
        from openpyxl.workbook.protection import WorkbookProtection
        wb.security = WorkbookProtection(lockStructure=True)  # impiedica re-afisarea sheetului
    except Exception:
        pass
    wb.save(client_path)
    print(f"Copie client: {client_path}")

# ===========================================================================
# MAIN
# ===========================================================================
if __name__=="__main__":
    folder = os.path.dirname(os.path.abspath(__file__))
    out_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(folder, "ARM_Cantitati_Structura.xlsx")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # sterge sheet-ul implicit

    note_tpl = citeste_note_template(folder)   # observatii preluate din template

    corpuri = []  # (eticheta, sheet_name, tot_row) pentru centralizator
    for sheet_cfg in SHEETS:
        sname = sheet_cfg["name"]

        # Sheet de tip cantitativ (incinta + consolidare teren) — alt format
        if sheet_cfg.get("tip") == "cantitativ":
            print(f"  Cantitativ: {sname}")
            write_cantitativ(wb.create_sheet(title=sname))
            continue

        titlu = sheet_cfg["titlu"]
        ifc_f = sheet_cfg["ifc"]

        ws = wb.create_sheet(title=sname)
        apply_col_widths(ws)

        clase    = sheet_cfg.get("clase")
        indici   = sheet_cfg.get("indici")
        majorari = sheet_cfg.get("majorari")
        # Note sub tabel: armatura + observatii din template + specifice
        notes = []
        if sname == "Infrastructura":
            if note_tpl["infra_hidro"]:
                notes.append(note_tpl["infra_hidro"])
            notes.append(NOTA_ARMATURA)
            notes += note_tpl["infra"]
        elif sname.startswith("Suprastructura"):
            notes.append(NOTA_ARMATURA)
            notes += note_tpl["supra"]
            notes.append(NOTA_ACOPERIS)
        tot_row = None
        if ifc_f:
            ifc_path = os.path.join(folder, ifc_f)
            if os.path.exists(ifc_path):
                rezultate, storey_order, supr_nivel = process_ifc(ifc_path)
                # Aplica override suprafete (ex: Parter preia suprafata de la Parter dx)
                for nivel_src, nivel_ref in sheet_cfg.get("supr_inherit", {}).items():
                    if nivel_ref in supr_nivel:
                        supr_nivel[nivel_src] = supr_nivel[nivel_ref]
                        print(f"    Suprafata '{nivel_src}' -> preluata de la '{nivel_ref}': {supr_nivel[nivel_ref]:.0f} mp")
                tot_row = write_sheet(ws, titlu, rezultate, storey_order, supr_nivel,
                                      clase=clase, indici=indici, majorari=majorari, notes=notes)
            else:
                print(f"  IFC negasit: {ifc_f} — sheet gol")
                write_sheet(ws, titlu, None, [], {}, clase=clase, indici=indici)
        else:
            print(f"  Sheet gol: {sname}")
            write_sheet(ws, titlu, None, [], {}, clase=clase, indici=indici)

        if tot_row:
            corpuri.append((titlu.replace("SUPRASTRUCTURA — ", "").replace("INFRASTRUCTURA","Infrastructura"),
                            sname, tot_row))

    # Centralizator la final
    wsc = wb.create_sheet(title="Centralizator")
    write_centralizator(wsc, corpuri)

    wb.save(out_path)
    print(f"\nSalvat (master): {out_path}")

    # Copie pentru client (doar printabil, formule ascunse, fara centralizator)
    base, ext = os.path.splitext(out_path)
    client_path = base + "_CLIENT" + ext
    make_client_copy(out_path, client_path)
    print("Gata!")
